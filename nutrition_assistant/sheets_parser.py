"""
Utilities for reading recent gym sessions from Google Sheets or local files.

Primary source:
- Google Sheets via service account credentials in `google_credentials.json`

Fallback sources:
- `gym_history.xlsx`
- `gym_history.csv`
- `gym_history.json`

The parser is intentionally tolerant with column names so it can work with
existing sheets without forcing a rigid schema.
"""

from __future__ import annotations

import csv
import json
import os
import re
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

try:
    import gspread
    from google.oauth2.service_account import Credentials
except Exception:  # pragma: no cover - optional dependency path
    gspread = None
    Credentials = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency path
    load_workbook = None


_DIR = Path(__file__).resolve().parent
_SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
_CONFIG_PATH = _DIR / "google_sheets_config.json"
_CREDENTIALS_PATH = _DIR / "google_credentials.json"
_LOCAL_XLSX_PATH = _DIR / "gym_history.xlsx"
_LOCAL_CSV_PATH = _DIR / "gym_history.csv"
_LOCAL_JSON_PATH = _DIR / "gym_history.json"


def get_recent_sessions(days: int = 7) -> tuple[list[dict[str, Any]], str]:
    """Return recent sessions plus source label: sheets, excel, or none."""
    normalized_days = max(1, int(days or 7))

    try:
        sessions = _load_google_sheets_sessions(normalized_days)
        if sessions:
            return sessions, "sheets"
    except Exception:
        # Fall through to local file fallback; endpoint should remain usable.
        pass

    for loader in (_load_excel_sessions, _load_csv_sessions, _load_json_sessions):
        try:
            sessions = loader(normalized_days)
            if sessions:
                return sessions, "excel"
        except Exception:
            continue

    return [], "none"


def _load_google_sheets_sessions(days: int) -> list[dict[str, Any]]:
    if gspread is None or Credentials is None:
        return []
    if not _CREDENTIALS_PATH.exists():
        return []

    config = _load_config()
    spreadsheet_id = (
        os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID")
        or config.get("spreadsheet_id")
        or config.get("sheet_id")
    )
    spreadsheet_name = (
        os.getenv("GOOGLE_SHEETS_SPREADSHEET_NAME")
        or config.get("spreadsheet_name")
    )
    if not spreadsheet_id and not spreadsheet_name:
        return []

    credentials = Credentials.from_service_account_file(
        str(_CREDENTIALS_PATH),
        scopes=_SCOPES,
    )
    client = gspread.authorize(credentials)

    if spreadsheet_id:
        spreadsheet = client.open_by_key(spreadsheet_id)
    else:
        spreadsheet = client.open(spreadsheet_name)

    rows: list[dict[str, Any]] = []
    for worksheet_name in _get_google_worksheet_candidates(config):
        try:
            worksheet = spreadsheet.worksheet(worksheet_name)
        except Exception:
            continue
        rows.extend(worksheet.get_all_records(default_blank=""))

    return _build_sessions(rows, days)


def _load_excel_sessions(days: int) -> list[dict[str, Any]]:
    if load_workbook is None:
        return []
    if not _LOCAL_XLSX_PATH.exists():
        return []

    config = _load_config()
    workbook = load_workbook(_LOCAL_XLSX_PATH, data_only=True, read_only=True)
    payload: list[dict[str, Any]] = []
    candidate_names = _get_excel_worksheet_candidates(config, workbook.sheetnames)
    for worksheet_name in candidate_names:
        if worksheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[worksheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        for row in rows[1:]:
            if not any(cell not in (None, "") for cell in row):
                continue
            payload.append({headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))})
    return _build_sessions(payload, days)


def _load_csv_sessions(days: int) -> list[dict[str, Any]]:
    if not _LOCAL_CSV_PATH.exists():
        return []

    with _LOCAL_CSV_PATH.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        return _build_sessions(list(reader), days)


def _load_json_sessions(days: int) -> list[dict[str, Any]]:
    if not _LOCAL_JSON_PATH.exists():
        return []

    with _LOCAL_JSON_PATH.open("r", encoding="utf-8") as fh:
        raw = json.load(fh)

    if isinstance(raw, dict):
        rows = raw.get("rows") or raw.get("sessions") or []
    else:
        rows = raw
    if not isinstance(rows, list):
        return []
    return _build_sessions(rows, days)


def _build_sessions(rows: list[dict[str, Any]], days: int) -> list[dict[str, Any]]:
    cutoff = date.today() - timedelta(days=max(0, days - 1))
    grouped: dict[tuple[str, str], dict[str, Any]] = {}

    for row in rows:
        if not isinstance(row, dict):
            continue
        normalized = {_normalize_key(k): v for k, v in row.items()}
        session_date = _extract_date(normalized)
        if not session_date or session_date < cutoff:
            continue

        exercise_name = _extract_str(normalized, _EXERCISE_KEYS)
        session_type = _extract_str(normalized, _TYPE_KEYS) or _infer_session_type(exercise_name)
        key = (session_date.isoformat(), session_type)

        if key not in grouped:
            grouped[key] = {
                "date": session_date,
                "type": session_type,
                "exercises": [],
            }

        exercise = _extract_exercise(normalized)
        if exercise is not None:
            grouped[key]["exercises"].append(exercise)

    sessions = sorted(
        grouped.values(),
        key=lambda item: item["date"] or date.min,
        reverse=True,
    )
    return sessions


def _extract_exercise(row: dict[str, Any]) -> dict[str, Any] | None:
    name = _extract_str(row, _EXERCISE_KEYS)
    if not name:
        return None

    kg_s1 = _extract_float(row, _KG_S1_KEYS)
    reps_s1 = _extract_int(row, _REPS_S1_KEYS)
    kg_s2 = _extract_float(row, _KG_S2_KEYS)
    reps_s2 = _extract_int(row, _REPS_S2_KEYS)
    volume = _extract_float(row, _VOLUME_KEYS)

    if volume <= 0:
        volume = kg_s1 * reps_s1 + kg_s2 * reps_s2

    return {
        "name": name,
        "volume": round(volume, 2),
        "kg_s1": round(kg_s1, 2),
        "reps_s1": reps_s1,
        "kg_s2": round(kg_s2, 2),
        "reps_s2": reps_s2,
        "compound": _extract_bool(row, _COMPOUND_KEYS),
    }


def _extract_date(row: dict[str, Any]) -> date | None:
    value = _extract_first(row, _DATE_KEYS)
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            # Excel serial date. Day 1 = 1899-12-31 with the usual adjustment.
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except Exception:
            return None

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _infer_session_type(exercise_name: str) -> str:
    if not exercise_name:
        return "Gym"
    lowered = exercise_name.lower()
    if any(token in lowered for token in ("bench", "press banca", "fondo", "hombro", "tricep")):
        return "Empuje"
    if any(token in lowered for token in ("remo", "dominada", "jalon", "bicep", "peso muerto")):
        return "Tiron"
    if any(token in lowered for token in ("sentadilla", "prensa", "cuadricep", "gluteo", "gemelo")):
        return "Piernas"
    return "Gym"


def _load_config() -> dict[str, Any]:
    if not _CONFIG_PATH.exists():
        return {}
    with _CONFIG_PATH.open("r", encoding="utf-8") as fh:
        raw = json.load(fh)
    return raw if isinstance(raw, dict) else {}


def _get_google_worksheet_candidates(config: dict[str, Any]) -> list[str]:
    explicit = os.getenv("GOOGLE_SHEETS_WORKSHEET") or config.get("worksheet")
    if explicit:
        return _unique_names([explicit])
    return _worksheet_candidates_from_date_window()


def _get_excel_worksheet_candidates(config: dict[str, Any], sheetnames: list[str]) -> list[str]:
    explicit = os.getenv("LOCAL_GYM_WORKSHEET") or config.get("local_worksheet") or config.get("worksheet")
    if explicit:
        return _unique_names([explicit])

    candidates = _worksheet_candidates_from_date_window()
    matches = [name for name in sheetnames if _normalize_key(name) in {_normalize_key(c) for c in candidates}]
    return _unique_names(matches or candidates or [sheetnames[0]])


def _worksheet_candidates_from_date_window() -> list[str]:
    today = date.today()
    previous = (today.replace(day=1) - timedelta(days=1))
    candidates: list[str] = []
    for current in (today, previous):
        year = current.year
        month_num = current.month
        month_num_2 = f"{month_num:02d}"
        month_es = _MONTH_NAMES_ES[month_num - 1]
        month_cap = month_es.capitalize()

        candidates.extend([
            month_es,
            month_cap,
            f"{month_num}",
            month_num_2,
            f"{year}-{month_num_2}",
            f"{month_num_2}-{year}",
            f"{year}_{month_num_2}",
            f"{month_num_2}_{year}",
            f"{month_es}_{year}",
            f"{month_cap}_{year}",
            f"{month_es} {year}",
            f"{month_cap} {year}",
            f"{month_num_2}_{month_es}",
            f"{month_num_2}_{month_cap}",
        ])

    return _unique_names(candidates + ["Entrenamientos"])


def _unique_names(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        normalized = _normalize_key(value)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        result.append(value)
    return result


def _extract_first(row: dict[str, Any], aliases: set[str]) -> Any:
    for key in aliases:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def _extract_str(row: dict[str, Any], aliases: set[str]) -> str:
    value = _extract_first(row, aliases)
    if value is None:
        return ""
    return str(value).strip()


def _extract_float(row: dict[str, Any], aliases: set[str]) -> float:
    value = _extract_first(row, aliases)
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else 0.0


def _extract_int(row: dict[str, Any], aliases: set[str]) -> int:
    return int(round(_extract_float(row, aliases)))


def _extract_bool(row: dict[str, Any], aliases: set[str]) -> bool:
    value = _extract_first(row, aliases)
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "si", "sí", "y", "x"}


def _normalize_key(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


_DATE_KEYS = {
    "date", "fecha", "dia", "day", "session_date", "workout_date",
}
_TYPE_KEYS = {
    "type", "tipo", "session_type", "workout_type", "rutina", "bloque", "day_type",
}
_EXERCISE_KEYS = {
    "exercise", "ejercicio", "name", "nombre", "movement", "movimiento",
}
_VOLUME_KEYS = {
    "volume", "volumen", "total_volume", "tonnage",
}
_COMPOUND_KEYS = {
    "compound", "compuesto", "multiarticular",
}
_KG_S1_KEYS = {
    "kg_s1", "peso_s1", "serie1_kg", "s1_kg", "set1_kg", "kg1", "peso1",
}
_REPS_S1_KEYS = {
    "reps_s1", "rep_s1", "serie1_reps", "s1_reps", "set1_reps", "reps1",
}
_KG_S2_KEYS = {
    "kg_s2", "peso_s2", "serie2_kg", "s2_kg", "set2_kg", "kg2", "peso2",
}
_REPS_S2_KEYS = {
    "reps_s2", "rep_s2", "serie2_reps", "s2_reps", "set2_reps", "reps2",
}

_MONTH_NAMES_ES = [
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
]
