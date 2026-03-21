"""
Parser para el Excel de entrenamiento de Alejandro.

Layout real (inspeccionado con índices de columna):
  col 2  → Nº ejercicio  |  'TIPO DE ENTRENAMIENTO'  |  'FECHA'
  col 3  → Nombre ejercicio  |  fecha (datetime)
  col 4  → Reps Serie 1  (float) o rango '12-15' (str, sin dato real)
  col 5  → Kg  Serie 1   (float)
  col 7  → Reps Serie 2  (float)
  col 8  → Kg  Serie 2   (float)
  col 10 → indicador de series ('2+1', 2.0 …)
  col 11 → rango rep / datetime (sin uso para cálculo)
"""

import datetime
import openpyxl
from pathlib import Path


EXCEL_PATH = Path("/mnt/c/Users/Alejandro/Desktop/entrenamiento/Alejandro - ENTRENAMIENTO.xlsx")

COMPOUND_KEYWORDS = [
    "peso muerto", "rdl", "sentadilla", "prensa", "leg press",
    "press plano", "press inclinado", "press 45", "press 30", "press 15",
    "press multipower", "kaz press", "press muy inclinado",
    "remo", "seal row", "jalón", "dominadas", "bulgarian", "zancada",
    "hip thrust", "rack pull",
]


def _is_compound(name: str) -> bool:
    n = name.lower()
    return any(kw in n for kw in COMPOUND_KEYWORDS)


def _safe_float(val) -> float | None:
    """Devuelve float si val es numérico y > 0, si no None."""
    if isinstance(val, (int, float)) and not isinstance(val, bool) and val > 0:
        return float(val)
    return None


def _parse_exercise_row(row: tuple) -> dict | None:
    """
    Intenta parsear una fila como ejercicio con datos reales.
    Devuelve None si la fila no corresponde a un ejercicio válido.
    """
    # Seguridad de longitud
    if len(row) < 9:
        return None

    idx  = _safe_float(row[2])   # índice ejercicio (1.0, 2.0, …)
    name = row[3]

    if idx is None or not isinstance(name, str) or not name.strip():
        return None

    reps1 = _safe_float(row[4])
    kg1   = _safe_float(row[5])
    reps2 = _safe_float(row[7])
    kg2   = _safe_float(row[8])

    # Si no hay kg reales (ejercicio solo con rangos, sin registrar), volumen = 0
    vol = 0.0
    if reps1 and kg1:
        vol += reps1 * kg1
    if reps2 and kg2:
        vol += reps2 * kg2

    return {
        "name":     name.strip(),
        "reps_s1":  reps1 or 0,
        "kg_s1":    kg1   or 0,
        "reps_s2":  reps2 or 0,
        "kg_s2":    kg2   or 0,
        "volume":   round(vol, 1),
        "compound": _is_compound(name),
    }


def parse_diary_sheet(sheet_name: str) -> list[dict]:
    """
    Devuelve lista de sesiones del sheet dado.
    Cada sesión: { type, date, exercises }
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    ws    = wb[sheet_name]
    rows  = list(ws.iter_rows(values_only=True))
    wb.close()

    sessions     = []
    i            = 0
    cur_type     = None
    cur_date     = None
    cur_exercises: list[dict] = []

    def _flush():
        nonlocal cur_type, cur_date, cur_exercises
        if cur_type and cur_exercises:
            sessions.append({
                "type":      cur_type,
                "date":      cur_date,
                "exercises": cur_exercises,
            })
        cur_type      = None
        cur_date      = None
        cur_exercises = []

    while i < len(rows):
        row = rows[i]
        # Seguridad de longitud
        padded = row + (None,) * max(0, 12 - len(row))

        marker = padded[2]

        # ── Inicio de nueva sesión ──────────────────────────────────────────
        if marker == "TIPO DE ENTRENAMIENTO":
            _flush()
            # El tipo está en col 4
            cur_type = str(padded[4]).strip() if padded[4] else "?"
            i += 1
            continue

        # ── Fila de fecha ───────────────────────────────────────────────────
        if marker == "FECHA":
            raw = padded[3]
            if isinstance(raw, datetime.datetime):
                cur_date = raw.date()
            i += 1
            continue

        # ── Cabecera de ejercicios (ignorar) ────────────────────────────────
        if isinstance(marker, str) and marker.startswith("ENTRENAMIENTO"):
            i += 1
            continue

        # ── Fila de ejercicio ───────────────────────────────────────────────
        ex = _parse_exercise_row(padded)
        if ex:
            cur_exercises.append(ex)

        i += 1

    _flush()
    return sessions


def get_all_diary_sheets() -> list[str]:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheets = [s for s in wb.sheetnames if s.startswith("DIARIO")]
    wb.close()
    return sheets


def get_current_month_sheet() -> str | None:
    MONTHS_ES = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
    }
    now  = datetime.datetime.now()
    name = f"DIARIO ALEX {MONTHS_ES[now.month]}"

    wb     = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    return name if name in sheets else None
